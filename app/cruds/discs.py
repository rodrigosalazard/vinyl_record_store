import os
from sqlalchemy.orm import Session
from sqlalchemy.sql.expression import or_
from passlib.hash import bcrypt
from datetime import datetime
from fastapi import HTTPException, status, File, UploadFile
from typing import List, Union, Optional
from email_validator import validate_email, EmailNotValidError
import json
from fastapi_pagination import paginate, Params
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image
import xlsxwriter

from schemas.Disc import DiscCreate, DiscUpdate
from cruds import files
from models.Disc import Disc
from models.File import File as FileModel
from external_api import connection


def get_discs(db: Session, skip: int = 0, limit: int = 200):
    """
    Regresa el listado de los discos de la base de datos.
    """
    return db.query(Disc).filter(Disc.deleted_at == None).offset(skip).limit(limit).all()

def get_disc(db: Session, disc_id: int):
    """
    Función busca un disco en la base de datos, siempre y cuando dicho disco no haya sido eliminado. Si lo encuentra, devuelve ese disco; de lo contrario, devuelve None.

    :param db: Conexión a la base de datos
    :type db: Session
    :param user_id: Identificador de disco.
    :type user_id: Int

    :return: disco de la base de datos
    :rtype: models.User

    """
    return db.query(Disc).filter(Disc.id == disc_id,Disc.deleted_at == None).first()


def create_disc(db: Session,disc: DiscCreate,uploaded_file: UploadFile = File(...)):
    """
    Crea un nuevo disco, valida su correo y confirma su contraseña, guarda fecha y hora de creación.
    """
    db_disc = Disc(album = disc.album, artist = disc.artist, genre = disc.genre, year = disc.year, price = disc.price,created_at = datetime.now(), updated_at = datetime.now())
    for key, value in disc.dict().items():
        if value is None or value == "":
            setattr(db_disc, key, getattr(db_disc, key))
        elif key== "created_at" or key== "updated_at" :
            setattr(db_disc, key, datetime.now())   
        else:
            setattr(db_disc, key, value)
    db.add(db_disc)
    db.commit()
    db.refresh(db_disc)

    db_file = None
    if uploaded_file:
        file_create = files.uploadFile('albums_covers',uploaded_file)
        db_file = FileModel(file_type = file_create.file_type, file_name = file_create.file_name, file_size = file_create.file_size, file_path = file_create.file_path, file_extension = file_create.file_extension,created_at = datetime.now(), updated_at = datetime.now())
        for key, value in file_create.dict().items():
            if value is None or value == "":
                setattr(db_file, key, getattr(db_file, key))
            elif key== "created_at" or key== "updated_at" :
                setattr(db_file, key, datetime.now())   
            else:
                setattr(db_file, key, value)
        db.add(db_file)
        db.commit()
        db.refresh(db_file)
    if db_file:
        setattr(db_disc, "file_id", db_file.id)
        db.add(db_disc)
        db.commit()
        db.refresh(db_disc)    
    return db_disc



def get_disc_by_name_and_artist(db: Session, album: str, artist:str):
    """
    Función busca un disco en la base de datos por album y artista.
    """
    return db.query(Disc).filter(Disc.album == album,Disc.artist == artist,Disc.deleted_at == None).first()


def delete_disc(db: Session, disc_id: int):
    """
    Elimina un disco, valida que exista y guarda fecha y hora de elminación.
    """
    db_disc = db.query(Disc).filter_by(id=disc_id).first()
    setattr(db_disc, 'deleted_at', datetime.now())     
    db.commit()
    db.refresh(db_disc)


def update_disc(db: Session, disc_id: int, disc: DiscUpdate, uploaded_file: UploadFile = File(...)):
    """
    Actualiza los campos del disco
    """
    db_disc = db.query(Disc).filter(Disc.id == disc_id, Disc.deleted_at == None).first()
    if not db_disc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="El disco no existe."
        ) 
    db_disc_same_album_artist = get_disc_by_name_and_artist(db=db,album=disc.album,artist=disc.artist)
    if db_disc_same_album_artist and db_disc_same_album_artist.id != disc_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, 
            detail="El album y artista ya existe."
        )
    db_disc = db.query(Disc).filter(Disc.id == disc_id, Disc.deleted_at == None).first()
    for key, value in disc.dict().items():
        if value is None or value == "":
            setattr(db_disc, key, getattr(db_disc, key))
        else:
            setattr(db_disc, key, value)
    setattr(db_disc, 'updated_at', datetime.now())
    db.add(db_disc)
    db.commit()
    db.refresh(db_disc)

    db_file = None
    if uploaded_file:
        file_create = files.uploadFile('albums_covers',uploaded_file)
        db_file = FileModel(file_type = file_create.file_type, file_name = file_create.file_name, file_size = file_create.file_size, file_path = file_create.file_path, file_extension = file_create.file_extension,created_at = datetime.now(), updated_at = datetime.now())
        for key, value in file_create.dict().items():
            if value is None or value == "":
                setattr(db_file, key, getattr(db_file, key))
            elif key== "created_at" or key== "updated_at" :
                setattr(db_file, key, datetime.now())   
            else:
                setattr(db_file, key, value)
        db.add(db_file)
        db.commit()
        db.refresh(db_file)
    if db_file:
        files.delete_file(db= db, file_id = db_disc.file_id)
        setattr(db_disc, "file_id", db_file.id)
        db.add(db_disc)
        db.commit()
        db.refresh(db_disc)   

    return db_disc


def search_disc(db: Session, page: int = 0, size: int = 10, search: Optional[str] = None):
    """
    Buscar discos por álbum o artista
    """
    # return db.query(Disc).filter(
    #     Disc.deleted_at == None,
    #     or_(Disc.album.ilike(f"%{search}%"), Disc.artist.ilike(f"%{search}%"))).offset(skip).limit(limit).all()
    query = db.query(Disc).filter(Disc.deleted_at == None)
    if search:
        query = query.filter(
            or_(Disc.album.ilike(f"%{search}%"), Disc.artist.ilike(f"%{search}%"))
        )
    total = query.count()
    results = query.all()

    # Crear el diccionario de respuesta
    return paginate(results)

def spotify_search_disc(db: Session,search: Optional[str] = None,size: Optional[int] = None):
    """
    Regresa el listado de los discos de la base de datos.
    """
    return connection.search_spotify_api(search,size)


def validate_headers(file: pd.ExcelFile, correct_headers: List[str]) -> Union[None, str]:
    sheet = file.sheet_names[0]  # asumimos que solo hay una hoja en el archivo
    df = file.parse(sheet)
    headers = list(df.columns)
    for header in headers:
        if header not in correct_headers:
            return header
    return None


def upload_discs(db: Session,uploaded_file: UploadFile = File(...)):
    file_create = files.uploadFile('excel',uploaded_file)
    workbook = openpyxl.load_workbook(file_create.file_path,)
    worksheet = workbook.active
    correct_headers = ["ARTISTA", "ALBUM", "GENERO", "AÑO", "PRECIO"]
    errors = []
    header_row = worksheet[1]
    actual_headers = [cell.value for cell in header_row]

    for header in correct_headers:
        if header not in actual_headers:
            errors.append(f"Falta cabecera o nombre incorrecto: {header}")
    if worksheet.max_row <= 1:
        errors.append("El archivo está vacío.")


    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{errors}")
    discs = []
    df = pd.read_excel(file_create.file_path, usecols=correct_headers, engine='openpyxl')
    for index, row in df.iterrows():
        try:
            float(row["PRECIO"])
        except ValueError:
            errors.append(f"Error en la línea:{index + 2} el valor PRECIO es incorrecto.")
        try:
            int(row["AÑO"])
        except ValueError:
            errors.append(f"Error en la línea:{index + 2} el valor AÑO es incorrecto.")

        db_disc = db.query(Disc).filter(Disc.album == row["ALBUM"],Disc.artist == row["ARTISTA"],Disc.deleted_at == None).first()
        if db_disc:
            setattr(db_disc, 'genre', row["GENERO"])
            setattr(db_disc, 'year', row["AÑO"])
            setattr(db_disc, 'price', row["PRECIO"])
            setattr(db_disc, 'updated_at', datetime.now())
            db.add(db_disc)
            db.commit()
        else:   
            disc = Disc(
                artist=row["ARTISTA"],
                album=row["ALBUM"],
                genre=row["GENERO"],
                year=row["AÑO"],
                price=row["PRECIO"],
                created_at= datetime.now(),
                updated_at= datetime.now(),
            )
            discs.append(disc)
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{errors}")
    db.add_all(discs)    
    db.commit()
    return {"message": "Discos cargados correctamente."}

def download_discs(db: Session, search: Optional[str] = None):
    query = db.query(Disc)
    if search:
        query = query.filter(
            or_(Disc.album.ilike(f"%{search}%"), Disc.artist.ilike(f"%{search}%"))
        )
    discs = query.all()
    # df = pd.DataFrame(
    #     [(a.artist, a.album, a.genre, a.year, f"${a.price:.2f}") for a in albums],
    #     columns=["Artist", "Album", "Genre", "Year", "Price"]
    # )

    # Crear una carpeta para guardar los archivos de exportación si no existe
    if not os.path.exists("exports"):
        os.makedirs("exports")

    # Crear el nombre del archivo de acuerdo al filtro aplicado
    timestamp = datetime.now().strftime('%d_%m_%Y_%H%M%S')
    filename = f"discs_{timestamp}"
    filepath = f"exports/excel/discs/{filename}.xlsx"

    # # Guardar el archivo de Excel
    # df.to_excel(filepath, index=False)

    # wb = openpyxl.load_workbook(filepath)
    # # Selecciona la hoja
    # sheet = wb.active

    # # Inserta la imagen en la celda
    # # img = Image('image.png')
    # # sheet.add_image(img, 'A1')

    # max_length = 0
    # for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
    #     for cell in row:
    #         try:
    #             if len(cell.value) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    # sheet.column_dimensions['A'].width = max_length
    # sheet.column_dimensions['B'].width = max_length

    # # Crear el workbook y la hoja
    workbook = xlsxwriter.Workbook(filepath)
    worksheet = workbook.add_worksheet('Lista de Discos')

    # Combinar las celdas A1 a A4
    worksheet.merge_range('A1:F5', '')

    img = Image.open('static/images/image.png')
    img_resized = img.resize((300, 300))
    img_width, img_height = img_resized.size
    worksheet.set_column('A:A', img_width/7)
    worksheet.set_row(0, img_height/15)
    worksheet.set_row(1, img_height/15)
    worksheet.set_row(2, img_height/15)
    worksheet.set_row(3, img_height/15)
    worksheet.insert_image('A1', 'static/images/image_logo.png')

    worksheet.set_column('A:A', 10)
    worksheet.set_column('B:B', 55)
    worksheet.set_column('C:C', 55)
    worksheet.set_column('D:D', 15)
    worksheet.set_column('E:E', 15)
    worksheet.set_column('F:F', 15)

    bold = workbook.add_format({'bg_color': '#009f86', 'align': 'center','color':"white"})
    worksheet.write('A6', 'ID', bold)
    worksheet.write('B6', 'ARTISA', bold)
    worksheet.write('C6', 'ALBUM', bold)
    worksheet.write('D6', 'GENERO', bold)
    worksheet.write('E6', 'AÑO', bold)
    worksheet.write('F6', 'PRECIO', bold)


    for idx, disc in enumerate(discs):
        row = idx + 7
        center = workbook.add_format({'align': 'center'})
        worksheet.write(f'A{row}', disc.id,center)
        worksheet.write(f'B{row}', disc.artist,center)
        worksheet.write(f'C{row}', disc.album)
        worksheet.write(f'D{row}', disc.genre,center)
        worksheet.write(f'E{row}', disc.year,center)
        currency_format = workbook.add_format({'num_format': '$#,##0.00','align': 'center'})
        worksheet.write(f'F{row}', disc.price, currency_format)
        worksheet.title = "Discos"

    workbook.close()

    return {"message": "Archivo excel generado correctamente.", "file_path": filepath}
